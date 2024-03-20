from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side

dict1 = {}
dict2 = {}

wb1_list = []
wb2_list = []

border = Border(left=Side(border_style='thin',color='000000'),
right=Side(border_style='thin',color='000000'),
top=Side(border_style='thin',color='000000'),
bottom=Side(border_style='thin',color='000000'))

def compare(wb1_list, wb2_list):    
    wb1_list = remove_whitespace_from_list(wb1_list)
    wb2_list = remove_whitespace_from_list(wb2_list)
    
    # print(wb1_list)
    # print(wb2_list)
    for q in range(len(wb1_list)):
        wb1 = load_workbook(wb1_list[q], read_only = False)
        wb2 = load_workbook(wb2_list[q], read_only = False)
    
        for xlsx in range(2):      
            if xlsx == 0:
                get_excel =  wb1
                wb_excel = wb1_list[q]
            else:
                get_excel =  wb2
                wb_excel = wb2_list[q]

            for sheet in range(len(get_excel.sheetnames)):
                work1 = wb1[wb1.sheetnames[sheet]]
                work2 = wb2[wb2.sheetnames[sheet]]

                make_dict(work1, work2)

                if xlsx == 0:
                    compare_a_to_b(work1, q)
                else:
                    compare_b_to_a(work2, q)  

                get_excel.save(f'maker_color_{wb_excel}')
                get_excel.close()
                dict1.clear()
                dict2.clear() 

wb1_count = []
def compare_a_to_b(work1, q):
    k = 1
    for key in dict1:
        for i in range(1, work1.max_row):
            # print(k, i, key, dict1[key][i-1])
            if dict1[key][i-1] is None and dict1[key][i-1] not in dict2.get(key):
                if key == 'Updated':
                    work1.cell(i+1 , k).font = Font(color="e06666")
                else:
                    work1.cell(i+1 , k).fill = PatternFill('solid',fgColor='e06666')
                    # print(work1.cell(i+1 , k).row)    #確認在哪一列
                    if work1.cell(i+1 , k).row not in wb1_count:
                        wb1_count.append(work1.cell(i+1 , k).row)
                    
        k+=1
    string_list = [str(element) for element in sorted(wb1_count)]
    result_string = ','.join(string_list)
    work1.cell(row=1, column=work1.max_column+1, value = 'Deleted').border  = border
    work1.cell(row=1, column=work1.max_column, value = 'Deleted').font  = Font(bold=True)
    work1.cell(row=2, column=work1.max_column, value = len(wb1_count)).font = Font(bold=True, color="e06666", size=15)
    work1.cell(row=2, column=work1.max_column+1, value = result_string)
    print(f'Finished comparing the {work1} in {wb1_list[q]}, total {len(wb1_count)} items deleted.')
    wb1_count.clear()
        
wb2_count = []        
def compare_b_to_a(work2, q):
    k = 1
    for key in dict2:
        for i in range(1, work2.max_row):
            # print(k, i, key, dict2[key][i-1])
            if dict2[key][i-1] is not None and dict2[key][i-1] not in dict1.get(key):
                if key == 'Updated':
                    work2.cell(i+1 , k).font = Font(color="a9d796")
                else:
                    work2.cell(i+1 , k).fill = PatternFill('solid',fgColor='a9d796')
                    if work2.cell(i+1 , k).row not in wb2_count:
                        wb2_count.append(work2.cell(i+1 , k).row)
                        
                
        k+=1
    string_list = [str(element) for element in sorted(wb2_count)]
    result_string = ','.join(string_list)
    work2.cell(row=1, column=work2.max_column+1, value = 'Added').border  = border
    work2.cell(row=1, column=work2.max_column, value = 'Added').font  = Font(bold=True)
    work2.cell(row=2, column=work2.max_column, value = len(wb2_count)).font = Font(bold=True, color="e06666", size=15)
    work2.cell(row=2, column=work2.max_column+1, value = result_string)
    print(f'Finished comparing the {work2} in {wb2_list[q]}, total {len(wb2_count)} items added.')
    wb2_count.clear()
    
def make_dict(work1, work2):
    for r in range(1, work1.max_column+1):
        key = work1.cell(1, r).value
        dict1[key] = []
        for c in range(2, work1.max_row+1):
            value = work1.cell(c, r).value
            dict1[key].append(value)
    
    # print(dict1)
    
    for r in range(1, work2.max_column+1):
        key = work2.cell(1, r).value
        dict2[key] = []
        for c in range(2, work2.max_row+1):
            value = work2.cell(c, r).value
            dict2[key].append(value)
            
    # print(dict2)
    
def remove_whitespace_from_list(input_list):
    return [item for item in input_list if item.strip()]
    
if __name__ == "__main__":
    while True:
        wb1_excel=input('Please input the first.xlsx file. Press Enter if you do not wish to continue the comparison: ')
        wb1_list.append(wb1_excel)
        wb2_excel=input('Please input the second.xlsx file. Press Enter if you do not wish to continue the comparison: : ')
        wb2_list.append(wb2_excel)
        if wb1_excel == '' and wb2_excel == '':
            break
    compare(wb1_list, wb2_list)